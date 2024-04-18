import requests
import json
import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Получаем данные за последние 24 часа
api_url = "https://min-api.cryptocompare.com/data/histohour?fsym=BTC&tsym=USDT&limit=24&toTs=" + str(int(datetime.datetime.now().timestamp())) + "&api_key=YOUR_API_KEY"
response = requests.get(api_url)
data = json.loads(response.text)['Data']

wb = openpyxl.load_workbook('btc_prices.xlsx')
sheet = wb.active
next_row = 1
while sheet['A' + str(next_row)].value is not None:
    next_row += 1

# Записываем данные в файл
for price_data in data:
    current_date = datetime.datetime.fromtimestamp(price_data['time']).strftime("%Y-%m-%d")
    current_time = datetime.datetime.fromtimestamp(price_data['time']).strftime("%H:00")
    btc_price_rounded = round(float(price_data['close']))

    sheet['A' + str(next_row)] = current_date
    sheet['B' + str(next_row)] = current_time
    sheet['C' + str(next_row)] = btc_price_rounded
    sheet['F' + str(next_row)] = current_time
    next_row += 1

if next_row > 2:
    for row in range(next_row - len(data), next_row):
        current_cell = sheet.cell(row=row, column=3)
        previous_cell = sheet.cell(row=row - 1, column=3)
        prev_prev_cell = sheet.cell(row=row - 2, column=3)

        if current_cell.value is not None and previous_cell.value is not None and prev_prev_cell.value is not None:
            if current_cell.value < previous_cell.value:
                current_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                percent_change = round((current_cell.value - previous_cell.value) / previous_cell.value * 100, 2)
                percent_change_str = "+" + str(percent_change) + "%" if percent_change >= 0 else str(percent_change) + "%"
                sheet.cell(row=row, column=4, value=percent_change_str)
                sheet.cell(row=row, column=4).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif current_cell.value > previous_cell.value:
                current_cell.fill = PatternFill(start_color="ccffcc", end_color="ccffcc", fill_type="solid")
                percent_change = round((current_cell.value - previous_cell.value) / previous_cell.value * 100, 2)
                percent_change_str = "+" + str(percent_change) + "%" if percent_change >= 0 else str(percent_change) + "%"
                sheet.cell(row=row, column=4, value=percent_change_str)
                sheet.cell(row=row, column=4).fill = PatternFill(start_color="8EFE03", end_color="8EFE03", fill_type="solid")

row = next_row - len(data) + 1
previous_sign = None
merged_range_start = None
merged_range_end = None
accumulated_change = 1.0
while row < next_row:
    current_cell_c = sheet.cell(row=row, column=3)
    current_cell_d = sheet.cell(row=row, column=4)

    if current_cell_d.value is not None:
        value = current_cell_d.value.replace('%', '')
        value = float(value)
        current_sign = 1 if value >= 0 else -1

        if current_sign != previous_sign or merged_range_start is None:
            if merged_range_start is not None:
                merged_range = merged_range_start + ":" + merged_range_end
                merged_cell = sheet[merged_range_start]
                overall_change = (accumulated_change - 1.0) * 100
                merged_cell.value = f"{'+' if overall_change >= 0 else ''}{overall_change:.2f}%"
                merged_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") if overall_change < 0 else PatternFill(start_color="8EFE03", end_color="8EFE03", fill_type="solid")
                merged_cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin', color='000000'),
                                                                    right=openpyxl.styles.borders.Side(style='thin', color='000000'),
                                                                    top=openpyxl.styles.borders.Side(style='thin', color='000000'),
                                                                    bottom=openpyxl.styles.borders.Side(style='thin', color='000000'))

            merged_range_start = get_column_letter(5) + str(row)
            merged_range_end = merged_range_start
            accumulated_change = 1.0

        else:
            merged_range_end = get_column_letter(5) + str(row)
        merged_range = merged_range_start + ":" + merged_range_end
        sheet.merge_cells(merged_range)
        accumulated_change *= (1 + value / 100)
        previous_sign = current_sign

    row += 1

if merged_range_start is not None:
    merged_range = merged_range_start + ":" + merged_range_end
    merged_cell = sheet[merged_range_start]
    overall_change = (accumulated_change - 1.0) * 100
    merged_cell.value = f"{'+' if overall_change >= 0 else ''}{overall_change:.2f}%"
    merged_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") if overall_change < 0 else PatternFill(start_color="8EFE03", end_color="8EFE03", fill_type="solid")
    merged_cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin', color='000000'),
                                                        right=openpyxl.styles.borders.Side(style='thin', color='000000'),
                                                        top=openpyxl.styles.borders.Side(style='thin', color='000000'),
                                                        bottom=openpyxl.styles.borders.Side(style='thin', color='000000'))

wb.save('btc_prices.xlsx')
print("New Bitcoin price successfully added to btc_prices.xlsx.")